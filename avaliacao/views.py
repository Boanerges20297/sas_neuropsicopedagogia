from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login as login_user, logout as logout_user
from django.contrib.auth.decorators import login_required
from django.contrib.auth import update_session_auth_hash
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.conf import settings
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_datetime
from django.db.models import Max, Sum
from datetime import datetime, timedelta
import requests
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .models import (
    PerfilUsuario, LogAuditoria, Paciente, AnotacaoAtendimento, AvaliacaoClinica,
    ConsultaIAClinica, FeedbackConsultaIA, ConsultaAtendimento
)
from .security_utils import decrypt_data, encrypt_data

# ==================== DECORADOR ADMIN ====================
def admin_required(view_func):
    """Garante acesso exclusivo para a Neuropsicopedagoga (administradora)."""
    def _wrapped_view_func(request, *args, **kwargs):
        if not request.user.is_authenticated or not request.user.is_admin():
            messages.error(request, 'Você não tem permissão para acessar esta área.')
            return redirect('index')
        return view_func(request, *args, **kwargs)
    return _wrapped_view_func


def staff_required(view_func):
    """Permite acesso operacional para administradores e usuários comuns."""
    def _wrapped_view_func(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if not (request.user.is_admin() or request.user.role == 'user'):
            messages.error(request, 'Você não tem permissão para acessar esta área.')
            return redirect('index')
        return view_func(request, *args, **kwargs)
    return _wrapped_view_func

# ==================== REGISTRO DE TRILHA DE AUDITORIA ====================
def register_audit_log(request, action):
    """Grava logs de auditoria automáticos na tabela MySQL para LGPD."""
    ip = request.META.get('HTTP_X_FORWARDED_FOR', request.META.get('REMOTE_ADDR', '127.0.0.1'))
    LogAuditoria.objects.create(
        user=request.user if request.user.is_authenticated else None,
        action=action,
        ip_address=ip
    )


def normalizar_data_br(valor):
    if not valor:
        return ""
    try:
        if "/" in valor:
            dt = datetime.strptime(valor, "%d/%m/%Y")
        else:
            dt = datetime.strptime(valor, "%Y-%m-%d")
        return dt.strftime("%d/%m/%Y")
    except Exception:
        return valor


def enviar_aprendizado_ia(content, context_domain="geral", clinical_category="casos_clinicos", target_age="todas"):
    if not content or len(content.strip()) < 10:
        return
    try:
        ai_learn_url = f"{settings.AI_SERVICE_URL}/api/v1/learn"
        payload = {
            "content": content.strip(),
            "context_domain": context_domain,
            "clinical_category": clinical_category,
            "target_age": target_age
        }
        requests.post(ai_learn_url, json=payload, timeout=2.5)
    except Exception:
        pass

# ==================== ROTAS DE AUTENTICAÇÃO ====================
def login_view(request):
    """Login que preserva o layout original marrom/dourado e valida cookies seguros."""
    if request.user.is_authenticated:
        return redirect('dashboard')

    if request.method == 'POST':
        email = request.POST.get('email')
        senha = request.POST.get('senha')
        remember = request.POST.get('remember') == 'on'

        try:
            user = PerfilUsuario.objects.get(email=email)
            authenticated_user = authenticate(request, username=user.username, password=senha)
        except PerfilUsuario.DoesNotExist:
            authenticated_user = None

        if authenticated_user:
            login_user(request, authenticated_user)
            if not remember:
                request.session.set_expiry(0) # Expira ao fechar navegador
            
            register_audit_log(request, "Realizou login no sistema.")
            messages.success(request, f"Bem-vinda, {authenticated_user.first_name or authenticated_user.username}!")
            return redirect('dashboard')
        else:
            messages.error(request, "E-mail ou senha incorretos.")

    return render(request, 'login.html')

def register_view(request):
    """Cadastro de novos usuários com validação forte de senhas."""
    if request.user.is_authenticated:
        return redirect('dashboard' if request.user.is_admin() else 'user_area')

    if request.method == 'POST':
        nome = request.POST.get('nome')
        email = request.POST.get('email')
        senha = request.POST.get('senha')
        data_nascimento = request.POST.get('data_nascimento')
        escolaridade = request.POST.get('escolaridade')
        telefone = request.POST.get('telefone')

        if PerfilUsuario.objects.filter(email=email).exists():
            messages.error(request, "Este e-mail já está cadastrado no sistema.")
            return render(request, 'register.html')

        if len(senha) < 8 or not any(c.isupper() for c in senha) or not any(c.isdigit() for c in senha):
            messages.error(request, "A senha deve ter pelo menos 8 caracteres, contendo pelo menos uma letra maiúscula e um número.")
            return render(request, 'register.html')

        username = email.split('@')[0] + "_" + str(PerfilUsuario.objects.count() + 1)
        user = PerfilUsuario.objects.create_user(
            username=username,
            email=email,
            password=senha,
            first_name=nome,
            role='user',
            data_nascimento=data_nascimento,
            escolaridade=escolaridade,
            telefone=telefone
        )
        
        register_audit_log(request, f"Novo usuário cadastrado: {email}")
        messages.success(request, "Cadastro realizado com sucesso! Faça login para continuar.")
        return redirect('login')

    return render(request, 'register.html')

@login_required
def logout_view(request):
    """Logout seguro do usuário."""
    register_audit_log(request, "Realizou logout do sistema.")
    logout_user(request)
    messages.info(request, "Você saiu da sua conta.")
    return redirect('login')

# ==================== ROTAS PÚBLICAS E REDIRECIONAMENTOS ====================
def index(request):
    """Página inicial com redirecionamento contextual."""
    if request.user.is_authenticated:
        return redirect('dashboard')
    return redirect('login')

@login_required
def user_area(request):
    """Área exclusiva para usuários comuns cadastrados."""
    return redirect('dashboard')

# ==================== CONTROLLER ADMINISTRATIVO (DASHBOARD) ====================
@login_required
@staff_required
def dashboard(request):
    """Agrega estatísticas operacionais de alto nível com Django ORM para o Chart.js."""
    total_responses = AvaliacaoClinica.objects.count()
    total_pacientes = Paciente.objects.count()
    consultas_ativas = ConsultaAtendimento.objects.filter(status__in=['agendada', 'em_atendimento']).count()
    
    today = timezone.now().date()
    new_today = AvaliacaoClinica.objects.filter(timestamp__date=today).count()
    
    total_users = PerfilUsuario.objects.count()
    admin_count = PerfilUsuario.objects.filter(role='admin').count()
    
    recent_responses_raw = AvaliacaoClinica.objects.all().order_by('-timestamp')[:5]
    recent_responses = []
    for r in recent_responses_raw:
        r.decrypt_sensitive()
        recent_responses.append(r)

    daily_labels = []
    daily_values = []
    for i in range(7):
        d = today - timedelta(days=6-i)
        daily_labels.append(d.strftime('%d/%m'))
        count = AvaliacaoClinica.objects.filter(timestamp__date=d).count()
        daily_values.append(count)

    last_r = AvaliacaoClinica.objects.all().order_by('-timestamp').first()
    last_response_date = last_r.timestamp.strftime('%d/%m/%Y') if last_r else 'Nenhuma'
    last_response_time = last_r.timestamp.strftime('%H:%M') if last_r else ''

    stats = {
        'total_responses': total_responses,
        'new_today': new_today,
        'total_users': total_users,
        'total_pacientes': total_pacientes,
        'consultas_ativas': consultas_ativas,
        'admin_count': admin_count,
        'last_response_date': last_response_date,
        'last_response_time': last_response_time,
        'recent_responses': recent_responses,
        'daily_labels': json.dumps(daily_labels),
        'daily_values': json.dumps(daily_values),
    }

    register_audit_log(request, "Acessou o Dashboard Administrativo.")
    return render(request, 'dashboard.html', {'stats': stats})

@login_required
@admin_required
def admin_avaliacoes(request):
    """Lista todas as avaliações clínicas / anamneses preenchidas."""
    status = request.GET.get('status', 'todas')
    responses_raw = AvaliacaoClinica.objects.all().order_by('-timestamp')
    if status == 'pendentes':
        responses_raw = responses_raw.filter(score=0)
    elif status == 'pontuadas':
        responses_raw = responses_raw.exclude(score=0)

    responses = []
    for r in responses_raw:
        r.decrypt_sensitive()
        if r.paciente:
            r.paciente.decrypt_sensitive()
        responses.append(r)
        
    register_audit_log(request, "Acessou a lista de avaliações clínicas.")
    return render(request, 'admin_avaliacoes.html', {
        'avaliacoes': responses,
        'status_filter': status,
        'total_avaliacoes': AvaliacaoClinica.objects.count(),
        'avaliacoes_pendentes': AvaliacaoClinica.objects.filter(score=0).count(),
        'avaliacoes_pontuadas': AvaliacaoClinica.objects.exclude(score=0).count(),
    })

@login_required
@admin_required
def view_avaliacao(request, avaliacao_id):
    """Detalhes clínicos do prontuário, consultando o microsserviço de IA de Borda."""
    avaliacao = get_object_or_404(AvaliacaoClinica, id=avaliacao_id)
    avaliacao.decrypt_sensitive()
    
    register_audit_log(request, f"Visualizou prontuário de avaliação ID {avaliacao_id} ({avaliacao.nome}).")

    ai_data = None
    ai_error = False
    
    try:
        context_domain = "paciente_" + str(avaliacao.paciente_id if avaliacao.paciente_id else "geral")
        
        payload = {
            "observed_characteristics": avaliacao.observed_characteristics or "",
            "andou_anos": avaliacao.andou_anos or "",
            "andou_meses": avaliacao.andou_meses or "",
            "falou_anos": avaliacao.falou_anos or "",
            "falou_meses": avaliacao.falou_meses or "",
            "escrita_comecou_anos": avaliacao.escrita_comecou_anos or "",
            "escrita_comecou_meses": avaliacao.escrita_comecou_meses or "",
            "leitura_comecou_anos": avaliacao.leitura_comecou_anos or "",
            "leitura_comecou_meses": avaliacao.leitura_comecou_meses or "",
            "pratica_esporte": avaliacao.pratica_esporte or "",
            "qual_esporte": avaliacao.qual_esporte or "",
            "assunto_interesse": avaliacao.assunto_interesse or "",
            "disciplinas_dificuldade": avaliacao.disciplinas_dificuldade or "",
            "context_domain": context_domain
        }
        
        ai_url = f"{settings.AI_SERVICE_URL}/api/v1/analyze"
        ai_resp = requests.post(ai_url, json=payload, timeout=4.0)
        
        if ai_resp.status_code == 200:
            ai_data = ai_resp.json().get('data')
        else:
            ai_error = True
    except Exception as e:
        print(f"⚠ Erro na comunicação com Microsserviço de IA: {e}")
        ai_error = True

    fields_list = []
    for f in AvaliacaoClinica._meta.fields:
        if f.name not in ['id', 'paciente', 'timestamp', 'score', 'scored_by', 'scored_at', 'notes']:
            val = getattr(avaliacao, f.name)
            fields_list.append((f.verbose_name or f.name, val))

    context = {
        'response': avaliacao,  # mantido como 'response' no context para compatibilidade com templates
        'fields_list': fields_list,
        'ai_data': ai_data,
        'ai_error': ai_error,
        'ai_scores_json': json.dumps(ai_data.get('scores')) if ai_data else json.dumps({
            'intelectual': 0, 'criatividade': 0, 'lideranca': 0, 'artes': 0, 'psicomotora': 0
        })
    }
    
    return render(request, 'view_response.html', context)

@login_required
@admin_required
def score_avaliacao(request, avaliacao_id):
    """Salva a pontuação atribuída e anotações clínicas via AJAX."""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            score = data.get('score', 0)
            notes = data.get('notes', '')

            avaliacao = get_object_or_404(AvaliacaoClinica, id=avaliacao_id)
            avaliacao.score = int(score)
            avaliacao.notes = notes
            avaliacao.scored_by = request.user
            avaliacao.scored_at = timezone.now()
            avaliacao.save()
            enviar_aprendizado_ia(
                content=f"Avaliação pontuada com score {score}. Parecer clínico: {notes}",
                context_domain=f"paciente_{avaliacao.paciente_id}" if avaliacao.paciente_id else "geral",
                clinical_category="pareceres_clinicos"
            )

            register_audit_log(request, f"Pontuou prontuário de avaliação ID {avaliacao_id} com {score} pontos.")
            return JsonResponse({'success': True, 'message': 'Parecer e pontuação salvos com sucesso!'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    return JsonResponse({'success': False, 'message': 'Método inválido'}, status=400)

@login_required
@admin_required
def admin_users(request):
    """Lista usuários cadastrados."""
    users = PerfilUsuario.objects.all().order_by('-created_at')
    register_audit_log(request, "Visualizou lista geral de usuários cadastrados.")
    return render(request, 'admin_users.html', {'users': users})


@login_required
@admin_required
def admin_user_create(request):
    if request.method == 'POST':
        nome = request.POST.get('nome', '').strip()
        email = request.POST.get('email', '').strip()
        senha = request.POST.get('senha', '').strip()
        role = request.POST.get('role', 'user').strip()
        telefone = request.POST.get('telefone', '').strip()
        escolaridade = request.POST.get('escolaridade', '').strip()

        if not nome or not email or not senha:
            messages.error(request, "Informe nome, e-mail e senha inicial.")
            return redirect('admin_user_create')

        if role not in dict(PerfilUsuario.ROLE_CHOICES):
            messages.error(request, "Perfil informado é inválido.")
            return redirect('admin_user_create')

        if PerfilUsuario.objects.filter(email=email).exists():
            messages.error(request, "Este e-mail já está cadastrado.")
            return redirect('admin_user_create')

        if len(senha) < 8 or not any(c.isupper() for c in senha) or not any(c.isdigit() for c in senha):
            messages.error(request, "A senha deve ter pelo menos 8 caracteres, uma letra maiúscula e um número.")
            return redirect('admin_user_create')

        username_base = email.split('@')[0]
        username = username_base
        suffix = 1
        while PerfilUsuario.objects.filter(username=username).exists():
            suffix += 1
            username = f"{username_base}_{suffix}"

        user = PerfilUsuario.objects.create_user(
            username=username,
            email=email,
            password=senha,
            first_name=nome,
            role=role,
            telefone=telefone,
            escolaridade=escolaridade,
        )
        register_audit_log(request, f"Criou usuário administrativo ID {user.id}.")
        messages.success(request, "Usuário criado com sucesso.")
        return redirect('admin_users')

    return render(request, 'admin_user_form.html')


@login_required
@admin_required
def admin_user_toggle_active(request, user_id):
    if request.method != 'POST':
        return redirect('admin_users')

    target = get_object_or_404(PerfilUsuario, id=user_id)
    if target.id == request.user.id:
        messages.error(request, "Você não pode desativar o próprio acesso.")
        return redirect('admin_users')

    target.is_active = not target.is_active
    target.save(update_fields=['is_active'])
    status = "ativado" if target.is_active else "desativado"
    register_audit_log(request, f"Usuário ID {target.id} {status}.")
    messages.success(request, f"Usuário {status} com sucesso.")
    return redirect('admin_users')


@login_required
@admin_required
def admin_user_change_role(request, user_id):
    if request.method != 'POST':
        return redirect('admin_users')

    target = get_object_or_404(PerfilUsuario, id=user_id)
    new_role = request.POST.get('role', '').strip()
    if new_role not in dict(PerfilUsuario.ROLE_CHOICES):
        messages.error(request, "Perfil informado é inválido.")
        return redirect('admin_users')

    if target.id == request.user.id and new_role != 'admin':
        messages.error(request, "Você não pode remover o próprio perfil administrativo.")
        return redirect('admin_users')

    target.role = new_role
    target.save(update_fields=['role'])
    register_audit_log(request, f"Alterou perfil do usuário ID {target.id} para {new_role}.")
    messages.success(request, "Perfil do usuário atualizado.")
    return redirect('admin_users')


@login_required
@staff_required
def perfil_usuario(request):
    if request.method == 'POST':
        acao = request.POST.get('acao', '').strip()

        if acao == 'dados':
            request.user.first_name = request.POST.get('nome', '').strip()
            request.user.email = request.POST.get('email', '').strip()
            request.user.telefone = request.POST.get('telefone', '').strip()
            request.user.escolaridade = request.POST.get('escolaridade', '').strip()
            request.user.save(update_fields=['first_name', 'email', 'telefone', 'escolaridade'])
            register_audit_log(request, "Atualizou os próprios dados de usuário.")
            messages.success(request, "Seus dados foram atualizados.")
            return redirect('perfil_usuario')

        if acao == 'senha':
            senha_atual = request.POST.get('senha_atual', '')
            nova_senha = request.POST.get('nova_senha', '')
            confirmar_senha = request.POST.get('confirmar_senha', '')

            if not request.user.check_password(senha_atual):
                messages.error(request, "Senha atual incorreta.")
                return redirect('perfil_usuario')
            if nova_senha != confirmar_senha:
                messages.error(request, "A confirmação da senha não confere.")
                return redirect('perfil_usuario')
            if len(nova_senha) < 8 or not any(c.isupper() for c in nova_senha) or not any(c.isdigit() for c in nova_senha):
                messages.error(request, "A nova senha deve ter pelo menos 8 caracteres, uma letra maiúscula e um número.")
                return redirect('perfil_usuario')

            request.user.set_password(nova_senha)
            request.user.save(update_fields=['password'])
            update_session_auth_hash(request, request.user)
            register_audit_log(request, "Alterou a própria senha.")
            messages.success(request, "Senha atualizada com sucesso.")
            return redirect('perfil_usuario')

    return render(request, 'perfil_usuario.html')

# ==================== PACIENTES E ATENDIMENTOS ====================
@login_required
@staff_required
def pacientes_list(request):
    pacientes = []
    for paciente in Paciente.objects.all().order_by('-updated_at'):
        paciente.decrypt_sensitive()
        paciente.data_nascimento = normalizar_data_br(paciente.data_nascimento)
        pacientes.append(paciente)
    register_audit_log(request, "Acessou a área de pacientes.")
    return render(request, 'pacientes_list.html', {'pacientes': pacientes})

@login_required
@staff_required
def paciente_create(request):
    if request.method == 'POST':
        data_nascimento = normalizar_data_br(request.POST.get('data_nascimento', '').strip())
        paciente = Paciente.objects.create(
            nome=request.POST.get('nome', '').strip(),
            data_nascimento=data_nascimento,
            responsavel=request.POST.get('responsavel', '').strip(),
            telefone=request.POST.get('telefone', '').strip(),
            email=request.POST.get('email', '').strip(),
            cpf=request.POST.get('cpf', '').strip(),
            identidade=request.POST.get('identidade', '').strip(),
            endereco=request.POST.get('endereco', '').strip(),
            bairro=request.POST.get('bairro', '').strip(),
            cep=request.POST.get('cep', '').strip(),
            escola=request.POST.get('escola', '').strip(),
            serie=request.POST.get('serie', '').strip(),
            queixa_principal=request.POST.get('queixa_principal', '').strip(),
            observacoes=request.POST.get('observacoes', '').strip(),
        )
        register_audit_log(request, f"Cadastrou paciente ID {paciente.id}.")
        messages.success(request, "Paciente cadastrado com sucesso.")
        if request.user.is_admin():
            return redirect('paciente_detail', paciente_id=paciente.id)
        return redirect('pacientes_list')
    return render(request, 'paciente_form.html')


@login_required
@staff_required
def paciente_update(request, paciente_id):
    if request.method != 'POST':
        return redirect('pacientes_list')

    paciente = get_object_or_404(Paciente, id=paciente_id)
    paciente.nome = request.POST.get('nome', '').strip()
    paciente.data_nascimento = normalizar_data_br(request.POST.get('data_nascimento', '').strip())
    paciente.responsavel = request.POST.get('responsavel', '').strip()
    paciente.telefone = request.POST.get('telefone', '').strip()
    paciente.email = request.POST.get('email', '').strip()
    paciente.cpf = request.POST.get('cpf', '').strip()
    paciente.identidade = request.POST.get('identidade', '').strip()
    paciente.endereco = request.POST.get('endereco', '').strip()
    paciente.bairro = request.POST.get('bairro', '').strip()
    paciente.cep = request.POST.get('cep', '').strip()
    paciente.escola = request.POST.get('escola', '').strip()
    paciente.serie = request.POST.get('serie', '').strip()
    paciente.queixa_principal = request.POST.get('queixa_principal', '').strip()
    paciente.observacoes = request.POST.get('observacoes', '').strip()
    paciente.save()
    register_audit_log(request, f"Atualizou cadastro do cliente ID {paciente.id}.")
    messages.success(request, "Cliente atualizado com sucesso.")
    return redirect('pacientes_list')

@login_required
@admin_required
def paciente_detail(request, paciente_id):
    paciente = get_object_or_404(Paciente, id=paciente_id)
    paciente.decrypt_sensitive()
    paciente.data_nascimento = normalizar_data_br(paciente.data_nascimento)

    if request.method == 'POST':
        raw_data_consulta = request.POST.get('data_consulta', '').strip()
        tipo = request.POST.get('tipo', 'sessao')
        titulo = request.POST.get('titulo', '').strip()
        conteudo = request.POST.get('conteudo', '').strip()
        data_consulta = parse_date(raw_data_consulta) if raw_data_consulta else None

        if raw_data_consulta and not data_consulta:
            messages.error(request, "Data da consulta invalida.")
            return redirect('paciente_detail', paciente_id=paciente.id)

        if tipo not in dict(AnotacaoAtendimento.TIPO_CHOICES):
            messages.error(request, "Tipo de anotação inválido.")
            return redirect('paciente_detail', paciente_id=paciente.id)

        if not conteudo:
            messages.error(request, "Informe o conteúdo da anotação.")
            return redirect('paciente_detail', paciente_id=paciente.id)

        if len(titulo) > 160:
            messages.error(request, "O titulo deve ter no maximo 160 caracteres.")
            return redirect('paciente_detail', paciente_id=paciente.id)

        anotacao = AnotacaoAtendimento.objects.create(
            paciente=paciente,
            profissional=request.user,
            tipo=tipo,
            titulo=titulo,
            conteudo=conteudo,
            data_consulta=data_consulta,
        )
        enviar_aprendizado_ia(
            content=f"Anotação clínica ({tipo}). Título: {titulo or 'Sem título'}. Conteúdo: {conteudo}",
            context_domain=f"paciente_{paciente.id}",
            clinical_category="anotacoes_clinicas"
        )
        register_audit_log(request, f"Registrou anotação ID {anotacao.id} no paciente ID {paciente_id}.")
        messages.success(request, "Anotação registrada.")
        return redirect('paciente_detail', paciente_id=paciente.id)

    anotacoes = []
    for item in paciente.anotacoes.all().order_by('-created_at'):
        item.decrypt_sensitive()
        anotacoes.append(item)

    # Buscar avaliações vinculadas
    avaliacoes = paciente.avaliacoes.all().order_by('-timestamp')
    for av in avaliacoes:
        av.decrypt_sensitive()

    return render(request, 'paciente_detail.html', {
        'paciente': paciente, 
        'anotacoes': anotacoes,
        'avaliacoes': avaliacoes
    })

@login_required
@admin_required
def paciente_nova_anamnese(request, paciente_id):
    """Garante redirecionamento para ficha com ID do paciente pré-configurado."""
    return redirect(f"/anamnese/nova/?paciente_id={paciente_id}")

@login_required
@staff_required
def consultas_list(request):
    pacientes = []
    for paciente in Paciente.objects.all().order_by('-updated_at'):
        paciente.decrypt_sensitive()
        pacientes.append(paciente)

    if request.method == 'POST':
        paciente_id = request.POST.get('paciente_id', '').strip()
        data_hora_raw = request.POST.get('data_hora', '').strip()
        motivo = request.POST.get('motivo', '').strip()
        observacoes = request.POST.get('observacoes_agendamento', '').strip()

        paciente = Paciente.objects.filter(id=paciente_id).first() if paciente_id.isdigit() else None
        data_hora = parse_datetime(data_hora_raw) if data_hora_raw else None
        if data_hora and timezone.is_naive(data_hora):
            data_hora = timezone.make_aware(data_hora)

        if not paciente:
            messages.error(request, "Selecione um cliente valido para marcar a consulta.")
            return redirect('consultas_list')

        if not data_hora:
            messages.error(request, "Informe data e horário da consulta.")
            return redirect('consultas_list')

        consulta = ConsultaAtendimento.objects.create(
            paciente=paciente,
            profissional=request.user,
            data_hora=data_hora,
            motivo=motivo,
            observacoes_agendamento=observacoes,
        )
        register_audit_log(request, f"Marcou consulta ID {consulta.id} para paciente ID {paciente.id}.")
        messages.success(request, "Consulta marcada com sucesso.")
        return redirect('consultas_list')

    consultas = []
    for consulta in ConsultaAtendimento.objects.select_related('paciente', 'profissional').order_by('data_hora'):
        consulta.decrypt_sensitive()
        consulta.paciente.decrypt_sensitive()
        consultas.append(consulta)

    register_audit_log(request, "Acessou a área de marcação de consultas.")
    return render(request, 'consultas_list.html', {
        'pacientes': pacientes,
        'consultas': consultas,
    })


@login_required
@staff_required
def consulta_update(request, consulta_id):
    if request.method != 'POST':
        return redirect('consultas_list')

    consulta = get_object_or_404(ConsultaAtendimento, id=consulta_id)
    paciente_id = request.POST.get('paciente_id', '').strip()
    data_hora_raw = request.POST.get('data_hora', '').strip()
    motivo = request.POST.get('motivo', '').strip()
    observacoes = request.POST.get('observacoes_agendamento', '').strip()
    status = request.POST.get('status', consulta.status).strip()

    paciente = Paciente.objects.filter(id=paciente_id).first() if paciente_id.isdigit() else None
    data_hora = parse_datetime(data_hora_raw) if data_hora_raw else None
    if data_hora and timezone.is_naive(data_hora):
        data_hora = timezone.make_aware(data_hora)

    if not paciente:
        messages.error(request, "Selecione um cliente válido para a consulta.")
        return redirect('consultas_list')
    if not data_hora:
        messages.error(request, "Informe data e horário da consulta.")
        return redirect('consultas_list')
    if status not in dict(ConsultaAtendimento.STATUS_CHOICES):
        messages.error(request, "Status informado é inválido.")
        return redirect('consultas_list')

    consulta.paciente = paciente
    consulta.data_hora = data_hora
    consulta.motivo = motivo
    consulta.observacoes_agendamento = observacoes
    consulta.status = status
    consulta.save()
    register_audit_log(request, f"Atualizou consulta ID {consulta.id}.")
    messages.success(request, "Consulta atualizada com sucesso.")
    return redirect('consultas_list')


@login_required
@admin_required
def atendimento_list(request):
    consultas = []
    qs = ConsultaAtendimento.objects.select_related('paciente', 'profissional').filter(
        status__in=['agendada', 'em_atendimento']
    ).order_by('data_hora')
    for consulta in qs:
        consulta.decrypt_sensitive()
        consulta.paciente.decrypt_sensitive()
        consultas.append(consulta)

    register_audit_log(request, "Acessou a área de atendimento.")
    return render(request, 'atendimento_list.html', {'consultas': consultas})


@login_required
@admin_required
def atendimento_detail(request, consulta_id):
    consulta = get_object_or_404(
        ConsultaAtendimento.objects.select_related('paciente', 'profissional'),
        id=consulta_id
    )
    consulta.decrypt_sensitive()
    consulta.paciente.decrypt_sensitive()

    if request.method == 'POST':
        acao = request.POST.get('acao', '').strip()
        anotacoes = request.POST.get('anotacoes_profissional', '').strip()

        if acao == 'iniciar':
            consulta.status = 'em_atendimento'
            consulta.profissional = request.user
            consulta.save()
            register_audit_log(request, f"Iniciou consulta ID {consulta.id}.")
            messages.success(request, "Consulta sinalizada como em atendimento.")
            return redirect('atendimento_detail', consulta_id=consulta.id)

        if acao == 'salvar':
            consulta.anotacoes_profissional = anotacoes
            consulta.save()
            register_audit_log(request, f"Salvou anotações profissionais da consulta ID {consulta.id}.")
            messages.success(request, "Anotações salvas.")
            return redirect('atendimento_detail', consulta_id=consulta.id)

        if acao == 'finalizar':
            if not anotacoes:
                messages.error(request, "Registre as anotações profissionais antes de finalizar a consulta.")
                return redirect('atendimento_detail', consulta_id=consulta.id)

            motivo_consulta = consulta.motivo or 'Consulta finalizada'
            consulta.anotacoes_profissional = anotacoes
            consulta.status = 'finalizada'
            consulta.encerrada_em = timezone.now()
            consulta.save()

            AnotacaoAtendimento.objects.create(
                paciente=consulta.paciente,
                profissional=request.user,
                tipo='sessao',
                data_consulta=consulta.data_hora.date(),
                titulo=motivo_consulta,
                conteudo=anotacoes,
            )
            enviar_aprendizado_ia(
                content=f"Consulta finalizada. Motivo: {motivo_consulta}. Anotações: {anotacoes}",
                context_domain=f"paciente_{consulta.paciente_id}",
                clinical_category="atendimentos_finalizados"
            )
            register_audit_log(request, f"Finalizou e baixou consulta ID {consulta.id}.")
            messages.success(request, "Consulta finalizada e baixada do atendimento.")
            return redirect('atendimento_list')

    anotacoes = []
    for item in consulta.paciente.anotacoes.all().order_by('-created_at')[:8]:
        item.decrypt_sensitive()
        anotacoes.append(item)

    return render(request, 'atendimento_detail.html', {
        'consulta': consulta,
        'paciente': consulta.paciente,
        'anotacoes': anotacoes,
    })

@login_required
@admin_required
def ia_consulta(request):
    selected_paciente_id = request.GET.get('paciente', '')
    pacientes = []
    for paciente in Paciente.objects.all().order_by('-updated_at')[:50]:
        paciente.decrypt_sensitive()
        pacientes.append(paciente)

    resultado = None
    consulta_id = None
    erro = None
    engine_info = {
        'nome': 'Neuro-Diagnosis Edge AI',
        'tipo': 'Microsserviço local FastAPI',
        'modelo': 'Motor heurístico próprio + memória semântica DSM-5-TR',
        'nuvem': 'Não utiliza nuvem nem APIs externas (totalmente local)',
        'url': settings.AI_SERVICE_URL,
    }

    if request.method == 'POST':
        paciente_id = request.POST.get('paciente_id')
        pergunta = request.POST.get('pergunta', '').strip()
        contexto = request.POST.get('contexto', '').strip()
        paciente = Paciente.objects.filter(id=paciente_id).first() if paciente_id else None
        if paciente:
            paciente.decrypt_sensitive()

        payload = {
            'observed_characteristics': contexto or pergunta,
            'andou_anos': '',
            'andou_meses': '',
            'falou_anos': '',
            'falou_meses': '',
            'escrita_comecou_anos': '',
            'escrita_comecou_meses': '',
            'leitura_comecou_anos': '',
            'leitura_comecou_meses': '',
            'pratica_esporte': '',
            'qual_esporte': '',
            'assunto_interesse': pergunta,
            'disciplinas_dificuldade': paciente.queixa_principal if paciente else '',
            'context_domain': f"paciente_{paciente.id}" if paciente else 'consulta_geral',
        }
        try:
            ai_url = f"{settings.AI_SERVICE_URL}/api/v1/analyze"
            ai_resp = requests.post(ai_url, json=payload, timeout=8.0)
            ai_resp.raise_for_status()
            resultado = ai_resp.json().get('data')
            consulta = ConsultaIAClinica.objects.create(
                paciente=paciente,
                profissional=request.user,
                pergunta=pergunta,
                contexto=contexto,
                resultado_json=json.dumps(resultado, ensure_ascii=False)
            )
            consulta_id = consulta.id

            if paciente and pergunta:
                AnotacaoAtendimento.objects.create(
                    paciente=paciente,
                    profissional=request.user,
                    tipo='ia',
                    titulo='Consulta IA',
                    conteudo=f"Pergunta: {pergunta}\n\nResultado: {json.dumps(resultado, ensure_ascii=False)}",
                )

            enviar_aprendizado_ia(
                content=f"Consulta IA. Pergunta: {pergunta}. Contexto: {contexto}. Resultado: {json.dumps(resultado, ensure_ascii=False)}",
                context_domain=f"paciente_{paciente.id}" if paciente else "consulta_geral",
                clinical_category="consultas_ia"
            )
            register_audit_log(request, "Realizou consulta à IA.")
        except Exception as exc:
            erro = f"Não foi possível consultar a IA local: {exc}"

    consultas_recentes = list(ConsultaIAClinica.objects.select_related('paciente').order_by('-created_at')[:10])
    for consulta in consultas_recentes:
        consulta.decrypt_sensitive()

    score_ia = FeedbackConsultaIA.objects.aggregate(total=Sum('peso')).get('total') or 0

    return render(request, 'ia_consulta.html', {
        'pacientes': pacientes,
        'resultado': resultado,
        'consulta_id': consulta_id,
        'consultas_recentes': consultas_recentes,
        'score_ia': score_ia,
        'erro': erro,
        'engine_info': engine_info,
        'selected_paciente_id': selected_paciente_id,
    })


@login_required
@admin_required
def feedback_consulta_ia(request, consulta_id):
    if request.method != 'POST':
        return redirect('ia_consulta')

    consulta = get_object_or_404(ConsultaIAClinica, id=consulta_id)
    julgamento = request.POST.get('julgamento', '').strip().lower()
    comentario = request.POST.get('comentario', '').strip()
    mapa_pesos = {'acerto': 2, 'parcial': 1, 'erro': -2}

    if julgamento not in mapa_pesos:
        messages.error(request, "Julgamento inválido para feedback da IA.")
        return redirect('ia_consulta')

    peso = mapa_pesos[julgamento]
    FeedbackConsultaIA.objects.create(
        consulta=consulta,
        avaliador=request.user,
        julgamento=julgamento,
        peso=peso,
        comentario=comentario,
    )

    consulta.decrypt_sensitive()
    categoria = "feedback_positivo_ia" if peso > 0 else "feedback_negativo_ia"
    enviar_aprendizado_ia(
        content=(
            f"Feedback IA ({julgamento}) com peso {peso}. "
            f"Pergunta original: {consulta.pergunta}. "
            f"Comentário clínico da profissional: {comentario or 'Sem comentário.'}"
        ),
        context_domain=f"paciente_{consulta.paciente_id}" if consulta.paciente_id else "consulta_geral",
        clinical_category=categoria,
    )

    register_audit_log(request, f"Registrou feedback IA ({julgamento}) na consulta {consulta_id}.")
    messages.success(request, "Feedback da IA registrado e incorporado ao aprendizado clínico.")
    return redirect('ia_consulta')

# ==================== PREENCHIMENTO E SUBMISSÃO DA ANAMNESE ====================
def nova_anamnese(request):
    """Renderiza a ficha clínica estruturada de 105 campos (Anamnese)."""
    CHECKBOX_OPTIONS = [
        "Facilidade em processar informações, integrar experiências e emitir respostas apropriadas",
        "Aprendizagem rápida/fácil e com pouca repetição",
        "Pensador crítico; gosta de lidar com problemas abstratos/complexos e propor novas soluções",
        "Boa memória e facilidade para acumular conhecimento",
        "Habilidade de raciocínio lógico-matemático",
        "Vocabulário avançado para a idade; verbalmente fluente",
        "Capacidade de generalizar e transferir aprendizagem",
        "Percepções incomuns na resolução de problemas",
        "Facilidade e agilidade para produzir ideias",
        "Flexibilidade ou facilidade para pensar fora dos padrões",
        "Originalidade de pensamento",
        "Capacidade de resolver problemas de forma criativa e efetiva",
        "Abertura a novas experiências e disposição para correr riscos",
        "Vê relações entre ideias diversas",
        "Independência e autonomia de pensamento",
        "Apurado senso de humor",
        "Interesse constante por certos tópicos",
        "Tendência a iniciar suas próprias atividades",
        "Persistência na realização de tarefas de interesse",
        "Auto-imposição para atingir a perfeição",
        "Ocupa seu tempo de forma produtiva",
        "Concentra-se por período prolongado sem aborrecer-se",
        "Preferência por responsabilidade pessoal sobre sua produção",
        "Obstinação em procurar informações sobre tópicos de interesse"
    ]
    
    paciente_id = request.GET.get('paciente_id')
    paciente = None
    if paciente_id:
        paciente = get_object_or_404(Paciente, id=paciente_id)
        paciente.decrypt_sensitive()
        
    return render(request, 'form.html', {
        'checkbox_options': CHECKBOX_OPTIONS,
        'paciente': paciente
    })

def salvar_anamnese(request):
    """Processa e salva as submissões da anamnese clínica no banco de dados MySQL de forma segura."""
    if request.method == 'POST':
        data = {}
        for f in AvaliacaoClinica._meta.fields:
            name = f.name
            if name not in ['id', 'paciente', 'timestamp', 'score', 'scored_by', 'scored_at', 'notes']:
                if name == 'observed_characteristics':
                    vals = request.POST.getlist('observed_characteristics')
                    data[name] = ", ".join(vals)
                else:
                    data[name] = request.POST.get(name, '').strip()

        data['data_nascimento'] = normalizar_data_br(data.get('data_nascimento', ''))
        data['bairro_data'] = normalizar_data_br(data.get('bairro_data', ''))

        # Vincular paciente se informado
        paciente_id = request.POST.get('paciente_id')
        if paciente_id and paciente_id.isdigit():
            data['paciente_id'] = int(paciente_id)

        # Se não vier nome preenchido mas vier paciente_id, preencher com o nome do paciente
        if not data.get('nome') and paciente_id:
            p = Paciente.objects.filter(id=paciente_id).first()
            if p:
                p.decrypt_sensitive()
                data['nome'] = p.nome

        new_resp = AvaliacaoClinica.objects.create(**data)
        
        # Dispara aprendizado silencioso no microsserviço de IA (se ativo) sobre este caso
        try:
            ai_learn_url = f"{settings.AI_SERVICE_URL}/api/v1/learn"
            payload = {
                "content": f"Caso clínico com observações: {data.get('observacoes')}. Características notadas: {data.get('observed_characteristics')}",
                "context_domain": f"paciente_{new_resp.paciente_id}" if new_resp.paciente_id else "geral",
                "clinical_category": "casos_clinicos",
                "target_age": "infantil" if int(data.get('idade_anos') or 10) < 12 else "adolescente"
            }
            requests.post(ai_learn_url, json=payload, timeout=2.0)
        except Exception as e:
            print(f"Sem conexão para aprendizado na IA: {e}")

        ip = request.META.get('HTTP_X_FORWARDED_FOR', request.META.get('REMOTE_ADDR', '127.0.0.1'))
        LogAuditoria.objects.create(
            user=request.user if request.user.is_authenticated else None,
            action=f"Nova avaliação preenchida na internet e salva. ID do Prontuário: {new_resp.id}",
            ip_address=ip
        )

        return render(request, 'thankyou.html')

    return redirect('nova_anamnese')

# ==================== EXPORTAÇÃO COMPLETA PARA EXCEL ====================
@login_required
@admin_required
def export_excel(request):
    """Gera planilha Excel profissional e estilizada contendo todos os prontuários e respostas."""
    register_audit_log(request, "Exportou a base de dados de avaliações para planilha Excel (.xlsx).")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Avaliações Clínicas"

    ws.views.sheetView[0].showGridLines = True

    fill_header = PatternFill(start_color="315b61", end_color="2a4d52", fill_type="solid")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    
    fill_even = PatternFill(start_color="F9F8F6", end_color="F9F8F6", fill_type="solid")
    fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    font_data = Font(name="Segoe UI", size=10)
    
    thin_border = Border(
        left=Side(style='thin', color='E5E0D8'),
        right=Side(style='thin', color='E5E0D8'),
        top=Side(style='thin', color='E5E0D8'),
        bottom=Side(style='thin', color='E5E0D8')
    )

    headers = ["ID Avaliação", "Paciente Vinculado", "Data e Hora", "Pontuação Atribuída", "Parecer Clínico"]
    fields_to_export = []
    
    for f in AvaliacaoClinica._meta.fields:
        if f.name not in ['id', 'paciente', 'timestamp', 'score', 'scored_by', 'scored_at', 'notes']:
            headers.append(f.verbose_name or f.name.replace("_", " ").capitalize())
            fields_to_export.append(f.name)

    ws.append(headers)

    for cell in ws[1]:
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    ws.row_dimensions[1].height = 28

    responses = AvaliacaoClinica.objects.all().order_by('-timestamp')
    row_num = 2
    
    for resp in responses:
        resp.decrypt_sensitive()
        
        paciente_nome = "Não Vinculado"
        if resp.paciente:
            resp.paciente.decrypt_sensitive()
            paciente_nome = resp.paciente.nome

        row_data = [
            resp.id,
            paciente_nome,
            resp.timestamp.strftime('%d/%m/%Y %H:%M'),
            resp.score,
            resp.notes or "Sem parecer atribuído"
        ]
        
        for field_name in fields_to_export:
            row_data.append(getattr(resp, field_name) or "")

        ws.append(row_data)
        
        row_fill = fill_even if row_num % 2 == 0 else fill_odd
        for cell in ws[row_num]:
            cell.fill = row_fill
            cell.font = font_data
            cell.border = thin_border
            if cell.column in [1, 2, 3, 4]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
        ws.row_dimensions[row_num].height = 22
        row_num += 1

    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        
        for cell in col:
            val = str(cell.value or '')
            if len(val) > max_len:
                max_len = len(val)
                
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f"attachment; filename=avaliacoes_neuropsicopedagogia_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    wb.save(response)
    return response
